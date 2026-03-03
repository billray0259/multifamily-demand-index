"""
Excel workbook export module.

Produces a multi-sheet .xlsx workbook with formatted tables and a
transparent Index Calculation sheet where every z-score, weighted
contribution, and 0-100 rescaling is a live Excel formula.
"""

import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from typing import Optional, Dict

from config import OUTPUT_METRICS, INDEX_WEIGHTS_FULL, INDEX_WEIGHTS_COSTAR_ONLY


# ── Shared palette ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=11)
FORMULA_FONT = Font(name="Calibri", size=10, color="1F3864")
NOTE_FONT    = Font(name="Calibri", italic=True, size=9, color="595959")
LABEL_FONT   = Font(name="Calibri", bold=True, size=11)
ACCENT_FILL  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_FILL     = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
TIER_FILLS   = {
    "High Demand":     PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Moderate Demand": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Low Demand":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
THIN_BORDER = Border(
    left=Side(style="thin",   color="D9D9D9"),
    right=Side(style="thin",  color="D9D9D9"),
    top=Side(style="thin",    color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _hdr(ws, row: int, col: int, val: str):
    c = ws.cell(row=row, column=col, value=val)
    c.fill   = HEADER_FILL
    c.font   = HEADER_FONT
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c


def _val(ws, row: int, col: int, val, fmt: str = None, align: str = "center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = BODY_FONT
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal=align)
    if fmt:
        c.number_format = fmt
    return c


def _fml(ws, row: int, col: int, formula, fmt: str = None,
         bold: bool = False, color: str = None):
    c = ws.cell(row=row, column=col, value=formula)
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="center")
    if bold or color:
        c.font = Font(name="Calibri", size=11, bold=bold, color=color or "000000")
    else:
        c.font = FORMULA_FONT
    if fmt:
        c.number_format = fmt
    return c


def _auto_width(ws, min_w: int = 10, max_w: int = 30):
    for col_cells in ws.columns:
        cl = get_column_letter(col_cells[0].column)
        lens = [len(str(c.value)) for c in col_cells if c.value is not None]
        if lens:
            ws.column_dimensions[cl].width = min(max(max(lens), min_w), max_w) + 2


def _coerce(val):
    """Convert numpy scalars to Python-native types for openpyxl."""
    if isinstance(val, np.integer):  return int(val)
    if isinstance(val, np.floating): return None if np.isnan(val) else float(val)
    if isinstance(val, np.bool_):    return bool(val)
    return val


def _metric_fmt(col_name: str) -> Optional[str]:
    for name, fmt_str, _ in OUTPUT_METRICS:
        if name == col_name:
            if "%" in fmt_str:                         return "0.0%"
            if "$" in fmt_str and ".2f" in fmt_str:    return "$#,##0.00"
            if "$" in fmt_str:                         return "$#,##0"
            if ",0f" in fmt_str or ",.0f" in fmt_str:  return "#,##0"
    return None


_COMP_NICE = {
    "absorption_pct":    "Absorption %",
    "absorption":        "Absorption %",
    "occupancy":         "Occupancy %",
    "rent_growth":       "Rent Growth %",
    "population_growth": "Pop Growth %",
    "income_growth":     "Income Growth %",
    "employment_rate":   "Employment Rate %",
    "employment_growth": "Employment Growth %",
    "construction_pct":  "Construction %",
    "deliveries_pct":    "Deliveries %",
    "prior_vacancy":     "Prior-Year Vacancy %",
}


# ── Sheet builder: Index Calculation ─────────────────────────────────────────

def _build_calculation_sheet(ws, components: pd.DataFrame, use_census: bool) -> Dict:
    """
    Write raw values (source data) and all computed values as Excel formulas.
    Returns a dict of column/row positions so other sheets can reference cells.
    """
    weights   = INDEX_WEIGHTS_FULL if use_census else INDEX_WEIGHTS_COSTAR_ONLY
    comp_keys = list(weights.keys())
    n_comp    = len(comp_keys)
    n_markets = len(components)

    def cl(c): return get_column_letter(c)

    # ── Title ─────────────────────────────────────────────────────────
    t = ws.cell(row=1, column=1, value="Multifamily Demand Index — Calculation Engine")
    t.font = Font(name="Calibri", bold=True, size=14)

    # ── Weights table ──────────────────────────────────────────────────
    W_HDR = 3;  W_DATA = 4
    _hdr(ws, W_HDR, 1, "Component")
    _hdr(ws, W_HDR, 2, "Weight (signed)")
    _hdr(ws, W_HDR, 3, "Effect on demand")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24

    weight_addr: Dict[str, str] = {}
    for i, (key, w) in enumerate(weights.items()):
        r = W_DATA + i
        ws.cell(row=r, column=1,
                value=_COMP_NICE.get(key, key.replace("_", " ").title())).font = BODY_FONT
        wc = ws.cell(row=r, column=2, value=w)
        wc.number_format = "0.00"
        wc.font   = Font(name="Calibri", size=11, bold=True, color="1F4E79")
        wc.border = THIN_BORDER
        wc.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3,
                value="Demand driver (+)" if w > 0 else "Supply pressure (−)").font = BODY_FONT
        weight_addr[key] = f"$B${r}"

    # ── Calculation table ──────────────────────────────────────────────
    DATA_HDR      = W_DATA + n_comp + 1
    DATA_START    = DATA_HDR + 1
    DATA_END      = DATA_START + n_markets - 1

    RAW_START     = 2
    Z_START       = RAW_START + n_comp
    CONTRIB_START = Z_START + n_comp
    RAW_SCORE_COL = CONTRIB_START + n_comp
    INDEX_COL     = RAW_SCORE_COL + 1

    raw_col     = {key: cl(RAW_START     + i) for i, key in enumerate(comp_keys)}
    z_col       = {key: cl(Z_START       + i) for i, key in enumerate(comp_keys)}
    contrib_col = {key: cl(CONTRIB_START + i) for i, key in enumerate(comp_keys)}


    # Column headers
    _hdr(ws, DATA_HDR, 1, "Market")
    for i, key in enumerate(comp_keys):
        nice = _COMP_NICE.get(key, key.replace("_", " ").title())
        _hdr(ws, DATA_HDR, RAW_START     + i, f"{nice}\n(raw)")
        _hdr(ws, DATA_HDR, Z_START       + i, f"{nice}\n(z-score)")
        _hdr(ws, DATA_HDR, CONTRIB_START + i, f"{nice}\n(z × weight)")
    _hdr(ws, DATA_HDR, RAW_SCORE_COL, "Raw Score\n(Σ contributions)")
    _hdr(ws, DATA_HDR, INDEX_COL,     "Demand Index\n(0 – 100)")
    ws.row_dimensions[DATA_HDR].height = 36
    # ws.freeze_panes = ws.cell(row=DATA_START, column=2)

    market_to_row: Dict[str, int] = {}

    for mkt_idx, (_, mrow) in enumerate(components.iterrows()):
        r   = DATA_START + mkt_idx
        mkt = str(mrow["Market"])
        market_to_row[mkt] = r

        # Market name (value)
        mc = ws.cell(row=r, column=1, value=mkt)
        mc.font = BODY_FONT; mc.border = THIN_BORDER
        mc.alignment = Alignment(horizontal="left")

        # Raw values — source data written as Python values
        for i, key in enumerate(comp_keys):
            rv = _coerce(mrow.get(f"{key}_raw", np.nan))
            _fml(ws, r, RAW_START + i, rv, "0.00")

        # Z-score formulas
        for i, key in enumerate(comp_keys):
            rc = raw_col[key]
            f = (f"=IFERROR(({rc}{r}-AVERAGE(${rc}${DATA_START}:${rc}${DATA_END}))"
                 f"/STDEV(${rc}${DATA_START}:${rc}${DATA_END}),0)")
            _fml(ws, r, Z_START + i, f, "0.000")

        # Weighted contribution formulas
        for i, key in enumerate(comp_keys):
            _fml(ws, r, CONTRIB_START + i,
                 f"={cl(Z_START + i)}{r}*{weight_addr[key]}", "0.000")

        # Raw score = SUM of contributions
        _fml(ws, r, RAW_SCORE_COL,
             f"=SUM({cl(CONTRIB_START)}{r}:{cl(CONTRIB_START + n_comp - 1)}{r})", "0.000")

        # Demand Index 0-100
        rs = cl(RAW_SCORE_COL)
        f  = (f"=IFERROR((${rs}{r}-MIN(${rs}${DATA_START}:${rs}${DATA_END}))"
              f"/(MAX(${rs}${DATA_START}:${rs}${DATA_END})"
              f"-MIN(${rs}${DATA_START}:${rs}${DATA_END}))*100,50)")
        _fml(ws, r, INDEX_COL, f, "0.0", bold=True, color="1F4E79")

        # Alternating row tint
        if mkt_idx % 2 == 1:
            for ci in range(1, INDEX_COL + 1):
                cell = ws.cell(row=r, column=ci)
                rgb  = getattr(cell.fill.start_color, "rgb", "FFFFFFFF")
                if cell.fill.fill_type in (None, "none") or rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = ALT_FILL

    # Colour scale on Index column
    ws.conditional_formatting.add(
        f"{cl(INDEX_COL)}{DATA_START}:{cl(INDEX_COL)}{DATA_END}",
        ColorScaleRule(
            start_type="min",       start_color="F8696B",
            mid_type="percentile",  mid_value=50, mid_color="FFEB84",
            end_type="max",         end_color="63BE7B",
        ),
    )

    for i in range(n_comp):
        ws.column_dimensions[cl(RAW_START     + i)].width = 14
        ws.column_dimensions[cl(Z_START       + i)].width = 14
        ws.column_dimensions[cl(CONTRIB_START + i)].width = 14
    ws.column_dimensions[cl(RAW_SCORE_COL)].width = 16
    ws.column_dimensions[cl(INDEX_COL)].width = 18

    ws.cell(row=DATA_END + 2, column=1,
            value="Formulas: z-score = (raw − AVERAGE) / STDEV  |  "
                  "Index = (raw_score − MIN) / (MAX − MIN) × 100").font = NOTE_FONT

    return {
        "ws_name":       ws.title,
        "data_start":    DATA_START,
        "data_end":      DATA_END,
        "n_markets":     n_markets,
        "market_to_row": market_to_row,
        "index_col":     cl(INDEX_COL),
        "index_col_num": INDEX_COL,
        "raw_col":       raw_col,
        "z_col":         z_col,
        "contrib_col":   contrib_col,
        "comp_keys":     comp_keys,
    }


# ── Sheet builder: Market Rankings ───────────────────────────────────────────

def _build_rankings_sheet(ws, rankings: pd.DataFrame, calc_info: Dict):
    """
    Market Rankings.  CoStar source metrics are written as values.
    Demand Index (VLOOKUP), Rank (RANK), and Tier (IF) are live Excel formulas.
    """
    cn       = calc_info["ws_name"]
    n        = len(rankings)
    data_end = n + 1  # header=row 1, data=rows 2..n+1

    has_period    = "Period" in rankings.columns
    col_market    = 1
    col_period    = 2 if has_period else None
    col_index     = 3 if has_period else 2
    col_rank      = col_index + 1
    col_tier      = col_rank  + 1
    metrics_start = col_tier  + 1

    idx_ltr  = get_column_letter(col_index)
    tier_ltr = get_column_letter(col_tier)
    mkt_ltr  = get_column_letter(col_market)

    metric_cols = [m[0] for m in OUTPUT_METRICS if m[0] in rankings.columns]

    # ── Headers ───────────────────────────────────────────────────────
    _hdr(ws, 1, col_market, "Market")
    if col_period: _hdr(ws, 1, col_period, "Period")
    _hdr(ws, 1, col_index, "Demand Index")
    _hdr(ws, 1, col_rank,  "Rank")
    _hdr(ws, 1, col_tier,  "Tier")
    for i, mc in enumerate(metric_cols):
        _hdr(ws, 1, metrics_start + i, mc.replace("_", " "))

    # ── Data rows ─────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(rankings.iterrows()):
        r = ri + 2

        # Market — value
        _val(ws, r, col_market, str(row["Market"]), align="left")

        # Period — value
        if col_period:
            _val(ws, r, col_period, row.get("Period"))

        # Demand Index — VLOOKUP from Index Calculation
        vlookup = (f"=VLOOKUP({mkt_ltr}{r},"
                   f"'{cn}'!$A:${calc_info['index_col']},"
                   f"{calc_info['index_col_num']},0)")
        _fml(ws, r, col_index, vlookup, "0.0", bold=True, color="1F4E79")

        # Rank — RANK on Demand Index column (descending)
        _fml(ws, r, col_rank,
             f"=RANK({idx_ltr}{r},${idx_ltr}$2:${idx_ltr}${data_end},0)", "0")

        # Tier — IF on Demand Index column
        _fml(ws, r, col_tier,
             f'=IF({idx_ltr}{r}>=67,"High Demand",'
             f'IF({idx_ltr}{r}>=33,"Moderate Demand","Low Demand"))')

        # Raw CoStar metrics — source data values
        for i, mc in enumerate(metric_cols):
            _val(ws, r, metrics_start + i, _coerce(row.get(mc)), _metric_fmt(mc))

    # ── Conditional formatting ─────────────────────────────────────────
    ws.conditional_formatting.add(
        f"{idx_ltr}2:{idx_ltr}{data_end}",
        ColorScaleRule(
            start_type="min",       start_color="F8696B",
            mid_type="percentile",  mid_value=50, mid_color="FFEB84",
            end_type="max",         end_color="63BE7B",
        ),
    )
    tier_range = f"{tier_ltr}2:{tier_ltr}{data_end}"
    for label, fill in TIER_FILLS.items():
        ws.conditional_formatting.add(tier_range,
            FormulaRule(formula=[f'{tier_ltr}2="{label}"'], fill=fill))

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet builder: Index Components ──────────────────────────────────────────

def _build_components_sheet(ws, components: pd.DataFrame,
                             calc_info: Dict, use_census: bool):
    """
    Index Components.  Every numeric cell is a direct cell reference to
    Index Calculation — no hard-coded computed values anywhere.
    """
    cn          = calc_info["ws_name"]
    weights     = INDEX_WEIGHTS_FULL if use_census else INDEX_WEIGHTS_COSTAR_ONLY
    comp_keys   = calc_info["comp_keys"]
    raw_col     = calc_info["raw_col"]
    z_col       = calc_info["z_col"]
    contrib_col = calc_info["contrib_col"]
    idx_col     = calc_info["index_col"]
    mkt_to_row  = calc_info["market_to_row"]

    hdrs = ["Market", "Demand Index"]
    for key in comp_keys:
        nice = _COMP_NICE.get(key, key.replace("_", " ").title())
        w    = weights[key]
        sign = "+" if w > 0 else "−"
        hdrs += [f"{nice}\n(raw)", f"{nice}\n(z-score)", f"{nice}\n({sign}{abs(w)*100:.0f}%)"]

    for i, h in enumerate(hdrs, 1):
        _hdr(ws, 1, i, h)
    ws.row_dimensions[1].height = 36

    for ri, (_, mrow) in enumerate(components.iterrows()):
        r   = ri + 2
        mkt = str(mrow["Market"])
        src = mkt_to_row.get(mkt)
        if src is None:
            continue

        _val(ws, r, 1, mkt, align="left")
        _fml(ws, r, 2, f"='{cn}'!${idx_col}${src}", "0.0", bold=True, color="1F4E79")

        c = 3
        for key in comp_keys:
            _fml(ws, r, c,     f"='{cn}'!${raw_col[key]}${src}",     "0.00");  c += 1
            _fml(ws, r, c,     f"='{cn}'!${z_col[key]}${src}",       "0.000"); c += 1
            _fml(ws, r, c,     f"='{cn}'!${contrib_col[key]}${src}", "0.000"); c += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Sheet builder: Census Demographics ───────────────────────────────────────

def _build_census_sheet(ws, census_snapshot: pd.DataFrame):
    census_cols = ["Market", "Year"]
    optional    = [
        "Population", "Median_Household_Income", "Employment_Rate",
        "Employment_Growth",
        "In_Migration", "In_Migration_Rate",
        "Population_Growth", "Median_Household_Income_Growth",
    ]
    census_cols += [c for c in optional if c in census_snapshot.columns]
    df = census_snapshot[census_cols].rename(columns={
        "Median_Household_Income":        "Median HH Income",
        "Employment_Rate":                "Employment Rate",
        "Employment_Growth":              "Employment Growth (YoY)",
        "In_Migration":                   "In-Migration",
        "In_Migration_Rate":              "In-Migration Rate",
        "Population_Growth":              "Population Growth (YoY)",
        "Median_Household_Income_Growth": "Income Growth (YoY)",
    })
    fmt_map = {
        "Employment Rate": "0.0%", "Employment Growth (YoY)": "0.0%",
        "In-Migration Rate": "0.0%",
        "Population Growth (YoY)": "0.0%", "Income Growth (YoY)": "0.0%",
        "Median HH Income": "$#,##0",
        "Population": "#,##0", "In-Migration": "#,##0",
    }
    for ci, col_name in enumerate(df.columns, 1):
        _hdr(ws, 1, ci, col_name)
    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 2
        for ci, col_name in enumerate(df.columns, 1):
            _val(ws, r, ci, _coerce(row[col_name]),
                 fmt_map.get(col_name),
                 align="left" if ci == 1 else "center")
    ws.freeze_panes = "A2"
    _auto_width(ws)


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_workbook(
    rankings: pd.DataFrame,
    components: pd.DataFrame,
    census_snapshot: Optional[pd.DataFrame] = None,
    use_census: bool = False,
) -> bytes:
    """
    Build a formula-driven .xlsx workbook and return it as bytes.

    Sheet order seen by the analyst:
      1. Market Rankings   — Demand Index / Rank / Tier are VLOOKUP/RANK/IF formulas
      2. Index Components  — every numeric cell references Index Calculation
      3. Census Demographics (if Census data available)
      4. Index Calculation — the engine: source data + all Excel formulas
    """
    wb = Workbook()

    # Build Index Calculation first (captures column/row positions)
    ws_calc       = wb.active
    ws_calc.title = "Index Calculation"
    calc_info = _build_calculation_sheet(ws_calc, components, use_census)

    # Build analyst-facing sheets
    ws_rank = wb.create_sheet("Market Rankings")
    _build_rankings_sheet(ws_rank, rankings, calc_info)

    ws_comp = wb.create_sheet("Index Components")
    _build_components_sheet(ws_comp, components, calc_info, use_census)

    if census_snapshot is not None and not census_snapshot.empty:
        ws_cen = wb.create_sheet("Census Demographics")
        _build_census_sheet(ws_cen, census_snapshot)

    # Reorder: analyst sheets first, engine last
    desired = ["Market Rankings", "Index Components"]
    if census_snapshot is not None and not census_snapshot.empty:
        desired.append("Census Demographics")
    desired.append("Index Calculation")
    wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)
    wb.active = wb["Market Rankings"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
